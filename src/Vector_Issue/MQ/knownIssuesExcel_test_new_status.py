import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Color, Alignment



class KnownIssuesExcel:
    """ Excel client class """

    def __init__(self, file):
        # initialize internal variables
        self.NUMBER = 2
        self.DATE = 3
        self.MODUL = 4
        self.FOUND_IN_VERSION = 5
        self.FIXED_IN_VERSION = 6
        self.TITLE = 7
        self.COLUMN_DESCRIPTION = 8
        self.COLUMN_ESCAN_REF = 9
        self.SOLUTION = 10
        self.MQSTATUS = 13

        self.file = file
        self.wb = openpyxl.load_workbook(file)
        self.sheetnames = self.wb.get_sheet_names()
        self.sheet = self.wb.get_sheet_by_name(self.sheetnames[0])
        self.defaultAlignment = openpyxl.styles.Alignment(horizontal='general',
                                vertical = 'top',
                                text_rotation = 0,
                                wrap_text = True,
                                shrink_to_fit = False,
                                indent = 0)
        self.sheet.column_dimensions['B'].auto_size = True
       

        #******************
        # validating
        #******************
        # Allowed Values are Open and Closed
        self.dvMQStatus = DataValidation(type="list", formula1='"Not analyzed,Not affected,Affected,Workaround used"', allow_blank=False)

        self.dvMQStatus.error = "Your entry is not in the list"
        self.dvMQStatus.errorTitle = 'Invalid Entry'
        
        self.sheet.add_data_validation(self.dvMQStatus)


    def removeExistingCellDataValidation(self, worksheet, cell):
        toRemove = []

        # Append all validation rules for cell to be removed.
        for validation in worksheet.data_validations.dataValidation:
            if validation.__contains__(cell):
                toRemove.append(validation)

        # Process all data validation rules set for removal.
        for rmValidation in toRemove:
            worksheet.data_validations.dataValidation.remove(rmValidation)

    def getAllAsDict(self):
        return self.getAllValuesInColumnsAsDict(self.COLUMN_ESCAN_REF, self.DATE, self.MODUL, 
                self.FOUND_IN_VERSION, self.FIXED_IN_VERSION, self.TITLE, self.COLUMN_DESCRIPTION, self.SOLUTION)

    def getEscanAndDescriptionAsDict(self):
        return self.getAllValuesInColumnsAsDict(self.COLUMN_ESCAN_REF, self.COLUMN_DESCRIPTION)


    def getAllValuesInColumnsAsDict(self, keycolumn=None, *columns):
        """
         Arugemnts: key:        key column index, the content of this column will be used as key in the dict
                                if key=None then the excel row number will be used as key
                    columns:    the columns specified here will be returned as list. if the key column is defined the
                                excel row index will be the first entry in the list
        """
        valuesInColumn = {}
        change_status = 0

        # go through all rows in the excelsheet
        for i in range(1, self.sheet.max_row+1):
            for column in columns:
                
                # Check if there is content in the keycolumn or in one of the passed columns
                if keycolumn != None and self.sheet.cell(row=i, column=keycolumn).value:
                
                    # Check if old status 'Open' or 'Closed' is present
                    if self.sheet.cell(row=8, column=13).value == "Open" or self.sheet.cell(row=8, column=13).value == "Closed":
                        change_status = 1
                       
                    # Remove old data validation and write as intitial status 'Not analyzed'   
                    if change_status == 1:
                        print("go")
                        column_str = str(i+1)
                        cell_status = "M"+column_str
                        self.removeExistingCellDataValidation(self.sheet, cell_status)
                        self.writeValueToCell(13, i+1, 'Not analyzed')
                    
                    # keycolumn will be used as index
                    valuesInColumn[self.sheet.cell(row=i, column=keycolumn).value] = [i]
                    for j in columns:
                        valuesInColumn[self.sheet.cell(row=i, column=keycolumn).value].append(self.sheet.cell(row=i, column=j).value)
                    continue
                if keycolumn == None and self.sheet.cell(row=i, column=column).value != None :
                    # if there is content then add it to the dictionary with the defined key column
                    # the index of the excel table is the first index in the dictionary if key is defined
                    # else the row will be used as key of the dict
                    # rowindex will be used as index
                    for j in columns:
                        valuesInColumn[i] = (self.sheet.cell(row=i, column=j).value)
                    continue
        return valuesInColumn

    def writeValueToCell(self, column, row, value):
        self.sheet.cell(row=row, column=column).alignment = self.defaultAlignment
        self.sheet.cell(row=row, column=column).value = value
        if column is self.MQSTATUS:
            self.dvMQStatus.add(self.sheet.cell(row=row, column=column))
        self.wb.save(self.file)

    def writeHeadingValueToCell(self, column, row, value):
        self.sheet.cell(row=row, column=column).alignment = Alignment(horizontal='center', vertical='top', wrap_text = True)
        self.sheet.cell(row=row, column=column).font = Font(name='CorpoS', size=10, bold=True)
        self.sheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor="A6A6A6")
        self.sheet.cell(row=row, column=column).value = value
        self.wb.save(self.file)

    def addLine(self):
        pass

    def set_border(self, fromColumn, fromRow, toColumn, toRow, ws=None):
        # if no worksheet is set use the default sheet from object
        if ws == None:
            ws = self.sheet

        # define border style
        thin_border = openpyxl.styles.Border(left=openpyxl.styles.borders.Side(style='thin'),
                             right=openpyxl.styles.borders.Side(style='thin'),
                             top=openpyxl.styles.borders.Side(style='thin'),
                             bottom=openpyxl.styles.borders.Side(style='thin'))

        # create cell range
        cell_range = openpyxl.utils.get_column_letter(fromColumn) + str(fromRow) + ':' + \
            openpyxl.utils.get_column_letter(toColumn) + str(toRow)

        rows = ws[cell_range]

        for row in rows:
            row[0].border = thin_border
            row[-1].border = thin_border
        for c in rows[0]:
            c.border = thin_border
        for c in rows[-1]:
            c.border = thin_border

    def set_alignment(self):
        pass

    def style_range(self, cell_range, border=openpyxl.styles.Border(), fill=None, font=None, alignment=None, ws = None):
        """
        Apply styles to a range of cells as if they were a single cell.

        :param ws:  Excel worksheet instance
        :param range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border
        :param fill: An openpyxl PatternFill or GradientFill
        :param font: An openpyxl Font object
        """
        if ws == None:
            ws = self.sheet

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill
