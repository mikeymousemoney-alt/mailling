import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Color, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl import styles
from openpyxl.formatting.rule import Rule


#CBDs
cbds = ['CBD2000373', 'CBD2000374', 'CBDtest']
#cbds = ['CBD2000779', 'CBD2000865', 'CBD0800064', 'CBD0800280', 'CBD0900105', 'CBD0900135', 'CBD0900354', 'CBD0900376', 'CBD1000187', 'CBD1100071', 'CBD1100085', 'CBD1100096', 'CBD1100102', 'CBD1200285', 'CBD1200413', 'CBD1300128', 'CBD1300404', 'CBD1300405', 'CBD1300581', 'CBD1300669', 'CBD1400105', 'CBD1400620', 'CBD1400794', 'CBD1500052', 'CBD1500056', 'CBD1500057', 'CBD1500431', 'CBD1500432', 'CBD1500433', 'CBD1500760', 'CBD1500761', 'CBD1500884', 'CBD1600095', 'CBD1600100', 'CBD1600268', 'CBD1600392', 'CBD1600394', 'CBD1600489', 'CBD1600671', 'CBD1600734', 'CBD1600781', 'CBD1600788', 'CBD1700205', 'CBD1700227', 'CBD1700242', 'CBD1700341', 'CBD1700342', 'CBD1700343', 'CBD1700344', 'CBD1700346', 'CBD1700414', 'CBD1700533', 'CBD1700556', 'CBD1700732', 'CBD1700863', 'CBD1700866', 'CBD1800141', 'CBD1800352', 'CBD1800379', 'CBD1800728', 'CBD1800813', 'CBD1800883', 'CBD1800899', 'CBD1801020', 'CBD1900222', 'CBD1900224', 'CBD1900230', 'CBD1900614', 'CBD1900950', 'CBD1901095', 'CBD2000056', 'CBD2000062', 'CBD2000373', 'CBD2000374', 'CBD2000660']
#cbds without CBD1900137 und CBD1900138
for cbd in cbds:
    try:
        #file = "C:\\temp\\Vector_Issue_Reports\\Test\\CBD2000374\\KnownBugsList.xlsx"
        file = "C:\\temp\\Vector_Issue_Reports\\Test\\"+cbd+"\\KnownBugsList.xlsx"
        #file = "X:\\ASR_Team\\Vector\\Microsar-Packages\\"+cbd+"\\KnownBugsList.xlsx"
        wb = openpyxl.load_workbook(file)

        black_colour_font = '000000'
        black_font = styles.Font(bold=False, color=black_colour_font)

        notAnalyzed_colour = 'FFC7CE'
        red_fill = styles.PatternFill(start_color=notAnalyzed_colour, end_color=notAnalyzed_colour, fill_type='solid')
        rule = Rule(type='containsText', text='Not analyzed', stopIfTrue=False)
        rule.dxf = DifferentialStyle(font=black_font, border=None, fill=red_fill)
        ws = wb.active
        rule.formula = ['NOT(ISERROR(SEARCH("Not analyzed",M1)))']
        ws.conditional_formatting.add('M1:M9999', rule)

        notAffected_colour = '99FF66'
        notAffected_fill = styles.PatternFill(start_color=notAffected_colour, end_color=notAffected_colour, fill_type='solid')
        rule2 = Rule(type='containsText', text='Not affected', stopIfTrue=False)
        rule2.dxf = DifferentialStyle(font=black_font, border=None, fill=notAffected_fill)
        ws = wb.active
        rule2.formula = ['NOT(ISERROR(SEARCH("Not affected",M1)))']
        ws.conditional_formatting.add('M1:M9999', rule2)

        affected_colour = 'FF0000'
        affected_fill = styles.PatternFill(start_color=affected_colour, end_color=affected_colour, fill_type='solid')
        rule3 = Rule(type='containsText', text='Affected', stopIfTrue=False)
        rule3.dxf = DifferentialStyle(font=black_font, border=None, fill=affected_fill)
        ws = wb.active
        rule3.formula = ['NOT(ISERROR(SEARCH("Affected",M1)))']
        ws.conditional_formatting.add('M1:M9999', rule3)

        workaroundUsed_colour = '00B0F0'
        workaroundUsed_fill = styles.PatternFill(start_color=workaroundUsed_colour, end_color=workaroundUsed_colour, fill_type='solid')
        rule4 = Rule(type='containsText', text='Not affected', stopIfTrue=False)
        rule4.dxf = DifferentialStyle(font=black_font, border=None, fill=workaroundUsed_fill)
        ws = wb.active
        rule4.formula = ['NOT(ISERROR(SEARCH("Workaround used",M1)))']
        ws.conditional_formatting.add('M1:M9999', rule4)        

        wb.save(file)
    except:
        print(cbd)