########################################################################################################################
# Author:   Markus Rigling
# Date:     28 November 2019
#
# Purpose:  Check outlook mails on Vector issue reports and export the attached xml-file.
#
# v1.0:     Initial Version
# v1.1:     Added external Integrator request
# v1.2:     Added logging
# v1.3:     Edited scan of multiple CBD-numbers
# v1.4:     Bug fixes
########################################################################################################################


########################################################################################################################
# imports
########################################################################################################################
import os  # Miscellaneous operating system interfaces
import sys
import re  # Regular expression operations
import copy  # Shallow and deep copy operations
import datetime as dt  # Basic date and time types
import threading  # Thread-based parallelism
import openpyxl #xlrd  # reading data and formatting information from Excel files
import logging  # logging
from Vector_Issue import make_config
import Vector_Issue.genKnownBugsList as genKnownBugsList # generate KnownBugsList
import json # read data from json files
import pandas as pd #reading and writing excell tables
import Vector_Issue.Vector_Issue as Vector_Issue
from Vector_Issue.utils import set_test_mode, test_log
from Vector_Issue import jira1
from Vector_Issue import graph_access_certStore
from pathlib import Path 
import argparse
import base64


import tkinter as tk
from tkinter.ttk import * 

import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
########################################################################################################################
# constants
########################################################################################################################
# check Emails and save attachments
c_periodCheckMessages_s = 28800  # change to 28800 (=8 h)?
#c_vectorEmailSubject = 'Vector CBD.+ Report:.+'
#c_vectorEmailSubjectWg = 'WG: Vector CBD.+ Report:.+'
#c_vectorEmailSubjectFw = 'FW: Vector CBD.+ Report:.+'
c_vectorEmailSubject = None # will be defined by the config.json
#c_subjectVector = "Vector"
c_subjectVector = "Report for"
#c_subjectReport = "Report"
c_subjectReport = ": "
c_pathIssueReports = None # will be defined by the config.json
#c_pathIssueReports = 'C:\\temp\\Vector_issue_mails\\Issue_Reports_Test\\%s\\%s\\IssueReports'
c_pathIssuesXlsx = "\\..\\.."
c_pathUnprocessedIssues = "\\..\\..\\.."
xmlFileDir = ""
xmlFileDir = ""

# read Autosarprojects.xlsx file and send notification Email
c_autosarprojectsListDir = None # will be defined by the config.json
#c_autosarprojectsListDir = 'C:\\temp\\Vector_issue_mails\\Test_Autosarprojects.xlsx'
c_autosarprojectsSheet = 'Overview'
c_columnMicrosarPackage = 10
c_columnFblPackage = 13
c_columnHsmPackage = 14
c_columnBswIntegrator = 5
c_columnFblIntegrator = 6
c_columnExtIntegrator = 7
c_columnProjectNumber = 0
c_columnProjectStatus = 4
c_typeMicrosarPackage = 1
c_typeFblPackage = 2
c_typeHsmPackage = 3
c_issuesEmail = None # will be defined by the config.json
c_mailFolder = None # will be defined by the config.json
c_marquardtEmail = 'marquardt.com'
c_subjectEmail = "%s: Received new Vector %s Report: %s"
c_bodyEmailIntern = "Received new Issue Report from Vector for: %s \nComponent: %s \nVersion: %s \nDate: %s"
c_bodyEmailExtern = "Received new Issue Report from Vector for: \nComponent: %s \nVersion: %s \nDate: %s"
c_bodyEmailProblem = "WARNING: XML file and/or PDF file not available or XML file not valid! \n\nReceived new Issue Report from Vector for: \nComponent: %s \nVersion: %s \nDate: %s"

# error messages
c_storagePdfFile = "Storage of PDF file not possible!"
c_storageXmlFile = "Storage of XML file not possible!"
c_validationError = "An error occurred!"
c_validationBswIntegrator = "BSW Integrator name not correct or none existing; no Email will be sent!"
c_noProjectNumber = "No P-Nr"

# logging
c_logFileDir = None # will be defined by the config.json
#c_logFileDir = 'C:\\temp\\Vector_issue_mails\\Logs\\VectorIssueReports_%s.log'
test_mode = 1
test_mail = ""
test_name = ""
set_read = 0
########################################################################################################################
# global variables
########################################################################################################################
subjectCBDNumber = ""
subjectDate = ""
version = ""
noProblem = 1
bswIntegrator = ""
extIntegrator = ""
pdfAttached = ""
xmlAttached = ""
bswIntegrators = []
extIntegrators = []
fblIntegrators = []
projectNumber = ""
projectNumbers = []
packageType = ""
packageTypeFound = []




class Vector_IssueApp:
    """Palim

    Args:
        tk (class): Some decription
    """
    
    def __init__(self,startDate):
        """ Constructor from Vector_IssueApp. 
        
            Will call automatically if an instance was created
        """
        global start_date
        start_date = startDate
        #super().__init__( )
        #
        #self.title('MQ Vector_Issue')
        #self.geometry('200x100')
        #self.quit()
        Vector_IssueApp.main()

    ########################################################################################################################
    # Get start date from user
    ########################################################################################################################
    @staticmethod
    def get_start_date():
        while True:
            start_date_input = input("Please enter the start date (DDMMYYYY): ")
            try:
                # Validate and return the entered date
                return dt.datetime.strptime(start_date_input, "%d%m%Y").date()
            except ValueError:
                print("Invalid date format. Please use DDMMYYYY.")    

    ########################################################################################################################
    # is_frozen
    ########################################################################################################################
    """
    Detect if we are run by an .exe (e.g. PyInstaller) or via python directly

    Args:
        None

    Returns:
        boolean : true, .exe, false, python native
    """
    def is_frozen():
        return getattr(sys, 'frozen', False)

    ########################################################################################################################
    # main function
    ########################################################################################################################
    @staticmethod
    def main():
        logging.debug("Vector_IssueApp.main-->")
        test_log("Vector_Issue.main-->")
        global start_date 
        start_date = None
        # Build config.json from config.txt and get json filename/path
        filename = make_config.main(Vector_IssueApp.is_frozen())
        
        with open(filename, 'r') as config_file:
            config = json.load(config_file)

        if len(sys.argv) > 1:
            arg = sys.argv[1]
            try:
                start_date = dt.datetime.strptime(arg, "%d%m%Y").date()
            except ValueError:
                print("Invalid date format. Please use DDMMYYYY.")
                start_date = Vector_IssueApp.get_start_date()
        else:
            # Prompt for start date if not provided
            start_date = Vector_IssueApp.get_start_date()

        Vector_IssueApp.UpdateConstants(config)

        # run script once
        Vector_IssueApp.CheckNewMails(start_date)

        # initialize ticker
        #ticker = threading.Event()
        #while not ticker.wait(c_periodCheckMessages_s):
            #Vector_IssueApp.CheckNewMails()
        logging.debug("Vector_IssueApp.main<--")
        test_log("Vector_Issue.main<--")
    ########################################################################################################################
    # Updating config-dependent constants
    ########################################################################################################################
    def UpdateConstants(config):
        print("Updating constants")
        logging.debug("Vector_IssueApp.UpdateConstants-->")
        def validate_field(field_name, is_empty_allowed=False):
            if field_name not in config:
                logging.error(f"Key '{field_name}' is missing in the configuration file.")
                raise KeyError(f"Key '{field_name}' is missing in the configuration file.")
            value = config[field_name]
            if not is_empty_allowed and value.strip() == "":
                logging.error(f"Field '{field_name}' cannot be empty.")
                raise ValueError(f"Field '{field_name}' cannot be empty.")
            return value

        global test_mode
        global set_read
        global move_mails
        global c_ASR_Functionality_Deactivated
        global partition_Change
        global send_Mails
        global create_tickets
        global test_mail
        global test_name
        global c_mailFolder
        global c_mailFolder_staging
        global c_mailFolder_prod
        global c_folderRead
        global c_folder_processed_mail_others
        global c_vectorEmailSubject
        global c_vectorEmailSubjectWg
        global c_vectorEmailSubjectFw
        global c_pathIssueReports
        global c_autosarprojectsListDir
        global c_logFileDir
        global c_columnProjectNumber
        global c_typeMicrosarPackage
        global c_typeFblPackage
        global c_typeHsmPackage
        global c_columnProjectStatus
        global c_columnBswIntegrator
        global c_columnFblIntegrator
        global c_columnExtIntegrator
        global c_columnMicrosarPackage
        global c_columnFblPackage
        global c_columnHsmPackage
        global c_start_date
        global vvm_api_url
        global c_cyberSecManager
        global c_vulnManager
        global c_mia_processing
        global c_jira_project_key_staging
        global c_jira_project_key_prod
        global c_jira_project_key
        global c_vector_mail_address
        global c_filter_nosecurity
        global c_folder_nosecurity
        global c_issuesEmail
        global c_issuesEmail_staging
        global c_issuesEmail_prod
        global c_client_id
        global c_tenant_id
        global c_cert_thumbprint

        test_mode = int(validate_field('test_mode'))
        set_read = int(validate_field('set_read'))
        move_mails = int(validate_field('move_mails'))
        test_mail = validate_field('test_mail', is_empty_allowed=True)
        test_name = validate_field('test_name')
        partition_Change = int(validate_field('change_to_C_partition'))
        send_Mails = int(validate_field('send_mails'))
        create_tickets = int(validate_field('create_tickets'))
        c_issuesEmail_prod = validate_field('mailbox_address')
        c_issuesEmail_staging = validate_field('mailbox_address_staging')
        if test_mode == 1:
            c_issuesEmail = c_issuesEmail_staging
        else:
            c_issuesEmail = c_issuesEmail_prod
        c_mailFolder_staging = validate_field('mailbox_folder_staging')
        c_mailFolder_prod = validate_field('mailbox_folder_prod')
        if test_mode == 1:
            c_mailFolder = c_mailFolder_staging
        else:
            c_mailFolder = c_mailFolder_prod

        c_folderRead = validate_field('folder_processed_mail_vector')
        c_folder_processed_mail_others = validate_field('folder_processed_mail_others')
        c_ASR_Functionality_Deactivated = validate_field('ASR_Functionality_Deactivated')
        c_vectorEmailSubject = config['email_subject']
        c_vectorEmailSubjectWg = 'WG: ' + c_vectorEmailSubject
        c_vectorEmailSubjectFw = 'FW: ' + c_vectorEmailSubject
        c_pathIssueReports = validate_field('path_to_issue_report')
        c_autosarprojectsListDir = validate_field('autosar_projects_list')
        c_logFileDir = validate_field('path_to_log_file')
        c_columnProjectNumber = int(validate_field('column_projectnumber'))
        c_typeMicrosarPackage = validate_field('column_microsar_package_type')
        c_typeFblPackage = validate_field('column_fbl_package_type')
        c_typeHsmPackage = validate_field('column_hsm_package_type')
        c_columnProjectStatus = int(validate_field('column_project_status'))
        c_columnBswIntegrator = int(validate_field('column_bsw_integrator'))
        c_columnFblIntegrator = int(validate_field('column_fbl_integrator'))
        c_columnExtIntegrator = int(validate_field('column_extern_integrator'))
        c_columnMicrosarPackage = int(validate_field('column_microsar_package'))
        c_columnFblPackage = int(validate_field('column_fbl_package'))
        c_columnHsmPackage = int(validate_field('column_hsm_package'))
        c_start_date = start_date
        vvm_api_url_staging = validate_field('VVM_API_URL_STAGING')
        vvm_api_url_prod = validate_field('VVM_API_URL_PROD')
        c_cyberSecManager = validate_field('cybersecurity_manager_email')
        c_vulnManager = validate_field('vulnerability_manager_email')
        c_mia_processing = validate_field('config_mia_processing')
        c_jira_project_key_staging = validate_field('jira_project_key_staging')
        c_jira_project_key_prod = validate_field('jira_project_key_prod')
        if test_mode == 1:
            c_jira_project_key = c_jira_project_key_staging
        else:
            c_jira_project_key = c_jira_project_key_prod

        c_vector_mail_address = validate_field('vector_mail_address')
        c_folder_nosecurity = validate_field('folder_no_security')
        c_filter_nosecurity = config.get('MIA_FILTER_NOSECURITY', [])
        if not isinstance(c_filter_nosecurity, list):
            c_filter_nosecurity = validate_field('MIA_FILTER_NOSECURITY')

        c_cert_thumbprint = validate_field('cert_thumbprint')
        c_client_id = validate_field('client_id')
        c_tenant_id = validate_field('tenant_id')


        set_test_mode(test_mode)
        if test_mode == 1:
            print('Test Mode is activated')
            vvm_api_url = vvm_api_url_staging
        else:
            vvm_api_url = vvm_api_url_prod

        if partition_Change == 1:
            c_pathIssueReports = Vector_IssueApp.change_partition(c_pathIssueReports)
            c_logFileDir = Vector_IssueApp.change_partition(c_logFileDir)

            print("New path to log file:", c_pathIssueReports)
            print("New path to issue report:", c_logFileDir)
            
        test_log("Vector_IssueApp.UpdateConstants<--")
        logging.debug("Vector_IssueApp.UpdateConstants<--")

    ########################################################################################################################
    # Changing the Partition to C: for testing
    ########################################################################################################################
    def change_partition(path, new_partition='C:'):
        # pattern to match partition letters
        pattern = r'^[a-zA-Z]:'
    
        # Search for the partition in the path
        match = re.match(pattern, path)
        if match:
            # Replace the old partition with the new partition
            old_partition = match.group(0)
            return new_partition + path[len(old_partition):]

        logging.error("No partition found in the path: %s", path)
        raise ValueError(f"No partition found in the path: {path}")
    
    ########################################################################################################################
    # Get Sender Emailadress
    ########################################################################################################################
    def get_sender_email(message):
        sender_info = message.get("from", {}).get("emailAddress", {})
        sender_email = sender_info.get("address")
        return sender_email

    ########################################################################################################################
    # function to check new E-Mails with Vector issue report subject and save attached xml-file
    ########################################################################################################################
    def CheckNewMails(start_date):
        test_log("-->Vector_IssueApp.CheckNewMails")
        logging.debug("-->Vector_IssueApp.CheckNewMails")
        # variables
        global subjectCBDNumber
        global subjectDate
        global version
        global noProblem
        global pdfAttached
        global xmlAttached
        global xmlFileDir
        global subject

        # Batch size for processing
        batch_size = 100

        processed_issues = []
        unprocessed_issues = []
        unprocessed_mail = []
        non_standard_mail=[]
        epic_keys=[]

        # get outlook inbox messages
        print("Accessing mailbox:", c_issuesEmail)
        messages = graph_access_certStore.main(start_date, c_cert_thumbprint, c_client_id, c_tenant_id, c_issuesEmail)

        numberOfFilteredMsg = len(messages)
        print("Number of messages from the date", start_date, ":", numberOfFilteredMsg)
        logging.info(f"Number of messages from the date {start_date}: {numberOfFilteredMsg}")


        c_folderProcessed = graph_access_certStore.get_folder_id(c_folderRead, c_issuesEmail)
        c_folderProcessedOthers = graph_access_certStore.get_folder_id(c_folder_processed_mail_others, c_issuesEmail)
        c_folderID_Nosecurity = graph_access_certStore.get_folder_id(c_folder_nosecurity, c_issuesEmail)


        print("send_Mails: " + str(send_Mails))

        if (move_mails == 0 and set_read == 0):
            batch_size = 50

        numFilteredBySubject = 0
        #while numberOfFilteredMsg > 0:

        current_batch = []
        epic_key = 0
        # find messages with Vector issue report subject
        for i in range(min(batch_size, numberOfFilteredMsg)):
            print("Going for email #:", i)
            print("")
            message = messages[i]
            subject = message['subject']
            email = genKnownBugsList
            print("Sender: ", Vector_IssueApp.get_sender_email(message))
            print()
            print()
            print ("c_vectorEmailSubject: " +c_vectorEmailSubject)
            test_log("Uncleaned Subject: " + subject)
            if re.fullmatch(c_vectorEmailSubjectWg, subject) or re.fullmatch(c_vectorEmailSubjectFw, subject):
                subject = re.sub(r'^(WG: |FW: )', '', message['subject'])
                subject = subject.strip()
                test_log("Cleaned Subject: " + message['subject'])

            else:
                test_log("Subject does not match FW or WG pattern")

            if any(re.fullmatch(pattern, subject) for pattern in c_filter_nosecurity):
                        graph_access_certStore.move_message(message['id'], c_issuesEmail, c_folderID_Nosecurity)
                        print("Moved to folder: " + str(c_folder_nosecurity))
                        logging.info("Message moved to no security folder: ", str(c_folder_nosecurity))
                        continue
                    
            if re.fullmatch(c_vectorEmailSubject, subject) or c_vectorEmailSubject == '':
                print("Subject: ", subject)
                logging.info("Subject: %s" % subject)

                if not re.match(c_vector_mail_address, str(Vector_IssueApp.get_sender_email(message)).strip(),re.IGNORECASE):
                    test_log("Sender is not Vector ", str(Vector_IssueApp.get_sender_email(message)).strip(),"!=",c_vector_mail_address)
                    logging.info("Sender is not Vector")
             

                    if c_mia_processing == "All_Emails": 
                        logging.debug("MIA processing is set to All Emails, processing the email")
                        non_standard_mail.append(message)
                        if create_tickets == 1:
                            epic_key = jira1.main(message, c_jira_project_key)
                        else:
                            print("Ticket would have been created here, but create_tickets is set to 0")
                            epic_key = epic_key + 1

                        epic_keys.append(epic_key)

                        if send_Mails == 1:
                            messageToSend = {
                                "subject": "Re: "+subject, # Keep the subject line
                                "body": "Thank you for reaching out to Marquardt Product Security Incident Response Team. Your request reached us, we're on it!",
                                "sentOnBehalfOfName" : c_issuesEmail,
                                "replyTo": str(Vector_IssueApp.get_sender_email(message)).strip()
                            }
                            print(messageToSend['replyTo'])
                            graph_access_certStore.send_mail("Re"+subject, messageToSend['body'], [messageToSend['replyTo']], c_issuesEmail)
                            logging.info("Reply sent for subject: " + subject)
                            print("Reply sent for subject: " + subject)

                        if set_read == 1:
                            graph_access_certStore.mark_message(message["id"],c_issuesEmail, is_read=1)
                                
                        current_batch.append(message)
                        continue
                    else:
                        print (str(Vector_IssueApp.get_sender_email(message)) + " is not Vector: " + c_vector_mail_address)
                        logging.debug(str(Vector_IssueApp.get_sender_email(message)) + " is not Vector: " + c_vector_mail_address)
                        continue

                # extract CBD number from E-Mail subject
                subject2ndPart = subject.split(c_subjectVector)[1]
                subjectDate = subject2ndPart.split(": ")[1]
                #print("Subject date: ", subjectDate)
                subjectCBDNumber = subject2ndPart.split(c_subjectReport)[0]
                subjectCBDNumber = subjectCBDNumber.replace(" ", "")
                #print("Subject CBD number: ", subjectCBDNumber)

                # get Integrator
                Vector_IssueApp.ReadAutosarprojects()
                    
                # get attachments
                attachments = graph_access_certStore.get_email_attachments(c_issuesEmail, message["id"])
                numberOfAttachments = len(attachments)

                print("Number of attachments: ", numberOfAttachments)
                logging.info("Number of attachments: %s" % numberOfAttachments)

                for attachment in attachments:
                    # get and save pdf file
                    pdfFile = attachment["filename"]
                    if re.search('.+pdf', pdfFile):
                        print("PDF file: ", pdfFile)
                        logging.info("PDF file: %s" % pdfFile)

                        # get version from name of pdf file
                        if "Cybersecurity-related" in subject:
                            version = pdfFile.split('_')[3]
                        else:
                            version = pdfFile.split('_')[2]
                            #print("Version: ", version)

                        #only for BSW purpose, CS has no write access to ASR_Team folders
                        if not c_ASR_Functionality_Deactivated:
                            # create directory for Issue Reports if it doesn't exist already
                            if not os.path.exists(c_pathIssueReports % (subjectCBDNumber, version)):
                                os.makedirs(c_pathIssueReports % (subjectCBDNumber, version))
                            # save attachment
                            try:
                                pdfAttached = c_pathIssueReports % (subjectCBDNumber, version) + '\\' + pdfFile
                                attachment.SaveAsFile(pdfAttached)
                            except:
                                print(c_storagePdfFile)
                                logging.info("%s" % c_storagePdfFile)

                    # get and save xml file and start genKnownBugsList.py script
                    xmlFile = attachment["filename"]
                    xmlFileContent = attachment["content"]
                    if re.search('.+xml', xmlFile):
                        print("XML file: ", xmlFile)
                        logging.info("XML file: %s" % xmlFile)

                        if not c_ASR_Functionality_Deactivated:
                            # save attachment
                            try:
                                xmlAttached = c_pathIssueReports % (subjectCBDNumber, version) + '\\' + xmlFile
                                attachment.SaveAsFile(xmlAttached)
                            except:
                                print(c_storageXmlFile)
                                logging.info("%s" % c_storageXmlFile)
                        if extIntegrator == "":
                            # start genKnownBugsList.py script and transfer xml_File as parameter
                            try:
                                xmlFileDir = c_pathIssueReports % (subjectCBDNumber, version) + '\\' + xmlFile
                                print("XML file directory: ", xmlFileDir)
                                logging.info("XML file directory: %s" % xmlFileDir)
                                genKnownBugsList.main(xmlFileDir, xmlFileContent, vvm_api_url, c_autosarprojectsListDir,subject, c_ASR_Functionality_Deactivated)
                                noProblem = 1
                            except:
                                print(c_validationError)
                                logging.info(c_validationError)
                                noProblem = 0
                        else:
                            noProblem = 1

                        if (numberOfAttachments % 2) != 0:
                            noProblem = 0

                    # notify BSW Integrator and/or Ext Integrator
                if noProblem == 1:
                    Vector_IssueApp.Notify()
                    # set unread message to read
                    if set_read == 0:
                        graph_access_certStore.mark_message(message["id"],mailbox=c_issuesEmail, is_read=0)
                    else:
                        graph_access_certStore.mark_message(message["id"],mailbox=c_issuesEmail, is_read=1)

                    current_batch.append(message)
                    try:
                        processed_issues=[
                            projectNumbers[-1], subjectCBDNumber, version, subjectDate,
                            bswIntegrators[-1], fblIntegrators[-1], extIntegrators[-1], 'Yes'
                        ]
                    except IndexError as e:
                        print(f"IndexError: {e} - likely empty list issue.")
                        processed_issues=[
                            'Unknown', subjectCBDNumber, version, subjectDate,
                            'Unknown', 'Unknown', 'Unknown', 'Error accessing list elements'
                        ]    
                else:
                    try:
                        processed_issues=[
                            projectNumbers[-1], subjectCBDNumber, version, subjectDate,
                            bswIntegrators[-1], fblIntegrators[-1], extIntegrators[-1], 'Issue with attachments'
                        ]
                    except IndexError as e:
                        print(f"IndexError: {e} - likely empty list issue.")
                        processed_issues=[
                            'Unknown', subjectCBDNumber, version, subjectDate,
                            'Unknown', 'Unknown', 'Unknown', 'Error accessing list elements'
                        ]
                    #Only for ASR functionality, CS does not have write access to ASR_Team folders
                if not c_ASR_Functionality_Deactivated:
                    Vector_IssueApp.create_processed_issues_excel([processed_issues])
            else:
                if re.fullmatch(c_filter_nosecurity, subject):
                    message.move(c_folder_nosecurity)
                    print("Moved to folder: " + str(c_folder_nosecurity))
                    logging.info("Moved to folder: " + str(c_folder_nosecurity))
                    continue

                if not re.match(c_vector_mail_address, str(Vector_IssueApp.get_sender_email(message)).strip()):
                    test_log("Sender is not Vector")

                print("Subject: ", subject)
                numFilteredBySubject += 1
                try:
                    # extract CBD number from E-Mail subject
                    subject2ndPart = subject.split(c_subjectVector)[1]
                    subjectDate = subject2ndPart.split(": ")[1]
                    #print("Subject date: ", subjectDate)
                    subjectCBDNumber = subject2ndPart.split(c_subjectReport)[0]
                    subjectCBDNumber = subjectCBDNumber.replace(" ", "")
                    #print("Subject CBD number: ", subjectCBDNumber)
                    attachments = message.Attachments
                    for attachment in attachments:
                        # get and save pdf file
                        pdfFile = attachment.FileName
                        if re.search('.+pdf', pdfFile):
                            print("PDF file: ", pdfFile)
                            logging.info("PDF file: %s" % pdfFile)

                            # get version from name of pdf file
                            if "Cybersecurity-related" in subject:
                                version = pdfFile.split('_')[3]
                            else:
                                version = pdfFile.split('_')[2]
                            #print("Version: ", version)
                    unprocessed_mail.append([subject, str(Vector_IssueApp.get_sender_email(message)).strip()])
                    unprocessed_issues.append([
                        'Unknown', subjectCBDNumber, version, subjectDate,
                        'Unknown', 'Unknown', 'Unknown', 'Subject does not match pattern'
                    ])
                except:
                    print("Subject: \"" + subject + "\" does not match pattern and cant be worked with")

        if move_mails ==1:
            for msg in current_batch:
                if not re.match(c_vector_mail_address, str(Vector_IssueApp.get_sender_email(message)).strip(),re.IGNORECASE):
                    graph_access_certStore.move_message(msg["id"], c_issuesEmail, c_folderProcessedOthers)
                else:
                    graph_access_certStore.move_message(msg["id"], c_issuesEmail, c_folderProcessed)
                    

        Vector_IssueApp.send_summary_Mail(non_standard_mail, epic_keys)
        if not c_ASR_Functionality_Deactivated:
            Vector_IssueApp.create_unprocessed_issues_excel(unprocessed_issues)
        test_log(f"Sending summary of {len(unprocessed_issues)} unprocessed issues.")
        Vector_IssueApp.send_summary_email(unprocessed_issues, unprocessed_mail)
        print("Number of Filtered Messege due to different Subject: ", numFilteredBySubject)



        test_log("<-- Vector_IssueApp.CheckNewMails")
        logging.debug("<-- Vector_IssueApp.CheckNewMails")

    ########################################################################################################################
    # function to read Autosarprojects.xlsx to get Microsar package responsible
    ########################################################################################################################
    def ReadAutosarprojects():
        test_log("-->Vector_IssueApp.ReadAutosarprojects")
        logging.debug("-->Vector_IssueApp.ReadAutosarprojects")
        # variables
        global bswIntegrator
        global extIntegrator
        global bswIntegrators
        global extIntegrators
        global fblIntegrators
        global projectNumbers
        global packageType

        # reset arrays
        del bswIntegrators[:]
        del extIntegrators[:]
        del fblIntegrators[:]
        del projectNumbers[:]
        del packageTypeFound[:]

        # open Autosarprojects.xlsx
        excelWorkbook = openpyxl.load_workbook(c_autosarprojectsListDir)
        sheet = excelWorkbook[c_autosarprojectsSheet]

        # search for CBD number and get BSW Integrator
        for rowNum, row in enumerate(sheet.iter_rows(values_only=True)):
            # scan column Microsar Package
            rowValue = row[c_columnMicrosarPackage]
            packageType = c_typeMicrosarPackage
            Vector_IssueApp.GetIntegrators(sheet, rowNum, rowValue, packageType)

            # scan column Flashbootloader Package
            rowValue = row[c_columnFblPackage]
            packageType = c_typeFblPackage
            Vector_IssueApp.GetIntegrators(sheet, rowNum, rowValue, packageType)

            # scan column HSM Package
            rowValue = row[c_columnHsmPackage]
            packageType = c_typeHsmPackage
            Vector_IssueApp.GetIntegrators(sheet, rowNum, rowValue, packageType)
        test_log("<-- Vector_IssueApp.ReadAutosarprojects")
        logging.debug("<-- Vector_IssueApp.ReadAutosarprojects")

    ########################################################################################################################
    # function to get BSW and Ext Integrators
    ########################################################################################################################
    def GetIntegrators(sheet, rowNum, rowValue, packageType):
        # variables
        global bswIntegrator
        global extIntegrator
        global bswIntegrators
        global extIntegrators
        global fblIntegrators
        global projectNumber
        global projectNumbers
        global packageTypeFound

        # search BSW and Ext Integrators for subject CBD number and Ensure rowValue is not None before accessing it
        if rowValue and re.search('%s' % subjectCBDNumber, rowValue):
            projectStatus = sheet.cell(rowNum+1, c_columnProjectStatus+1).value
            projectInactive = 0
            if projectStatus == "Stopped":
                projectInactive = 1
            if projectStatus == "Closed":
                projectInactive = 1
            if projectInactive != 1:
                bswIntegrator = sheet.cell(rowNum+1, c_columnBswIntegrator+1).value
                projectNumber = sheet.cell(rowNum+1, c_columnProjectNumber+1).value
                if projectNumber == "":
                    projectNumber = c_noProjectNumber
                #print("%s - BSW Integrator: %s" % (projectNumber, bswIntegrator))
                if bswIntegrator != "":
                    #logging.info("%s - BSW Integrator: %s" % (projectNumber, bswIntegrator))
                    pass
                extIntegrator = sheet.cell(rowNum+1, c_columnExtIntegrator+1).value
                #print("%s - Ext Integrator: %s" % (projectNumber, extIntegrator))
                if extIntegrator != "":
                    #logging.info("%s - Ext Integrator: %s" % (projectNumber, extIntegrator))
                    pass

                bswIntegrators.append(sheet.cell(rowNum+1, c_columnBswIntegrator+1).value)
                print("BSW Integrators: ", bswIntegrators)
                extIntegrators.append(sheet.cell(rowNum+1, c_columnExtIntegrator+1).value)
                print("Ext Integrators: ", extIntegrators)
                fblIntegrators.append(sheet.cell(rowNum+1, c_columnFblIntegrator+1).value)
                print("Fbl Integrators: ", fblIntegrators)
                projectNumbers.append(sheet.cell(rowNum+1, c_columnProjectNumber+1).value)
                print("Project numbers: ", projectNumbers)
                packageTypeFound.append(packageType)
                print("Package type: "+str(packageType))
        else:
            bswIntegrator = ""
            extIntegrator = ""

    ########################################################################################################################
    # function to build the integrator's email
    ########################################################################################################################
    def BuildIntegratorEmail(integrator):
        # variables
        global subjectCBDNumber
        global version
        global noProblem
        global bswIntegrator
        global extIntegrator
        global pdfAttached

        # init variables
        integratorsFirstNames = [None] * len(integrator)
        integratorsLastNames = [None] * len(integrator)
        integratorsEmails = [None] * len(integrator)
        noneCorrectIntegratorName = [None] * len(integrator)

        # build BSW Integrators Email addresses
        for i, j, k, l, m, n in zip(integrator, range(len(integratorsFirstNames)), range(len(integratorsLastNames)),
                                    range(len(integratorsEmails)), range(len(noneCorrectIntegratorName)), range(len(projectNumbers))):
            try:
                i = i.lstrip()  # delete leading spaces
                i = re.sub(' +', ' ', i)  # delete multiple spaces
                #print("Integrator without leading and multiple spaces: ", i)

                if "@" in i:
                    integratorsEmails[l] = i
                    #print("BSW Integrator Email available: ", integratorsEmails[l])
                elif "," in i:
                    integratorsFirstNames[j] = i.split(' ')[1]
                    integratorsFirstNames[j] = integratorsFirstNames[j].replace(",", "")
                    #print("BSW Integrator First Name: ", integratorsFirstNames[j])
                    integratorsLastNames[k] = i.split(' ')[0]
                    integratorsLastNames[k] = integratorsLastNames[k].replace(",", "")
                    #print("BSW Integrator Last Name: ", integratorsLastNames[k])
                else:
                    integratorsFirstNames[j] = i.split(' ')[0]
                    #print("BSW Integrator First Name: ", integratorsFirstNames[j])
                    integratorsLastNames[k] = i.split(' ')[1]
                    #print("BSW Integrator Last Name: ", integratorsLastNames[k])

                if "@" not in i:
                    integratorsEmails[l] = "%s.%s@%s" % (integratorsFirstNames[j], integratorsLastNames[k], c_marquardtEmail)
                #print("%s - BSW Integrator E-Mail: %s" % (projectNumbers[n], integratorsEmails[l]))
                #logging.info("%s - BSW Integrator E-Mail: %s" % (projectNumbers[n], integratorsEmails[l]))
                noneCorrectIntegratorName[m] = 0

            except:
                noneCorrectIntegratorName[m] = 1
                integratorsEmails[l] = ""
                if projectNumbers[n] == "":
                    projectNumbers[n] = c_noProjectNumber
                #print("%s - %s" % (projectNumbers[n], c_validationBswIntegrator))
                #logging.info("%s - %s" % (projectNumbers[n], c_validationBswIntegrator))

        return integratorsEmails, noneCorrectIntegratorName
    ########################################################################################################################
    # function to send E-Mail to BSW and/or Ext Integrator
    ########################################################################################################################
    def Notify():
        logging.debug("--> Vector_IssueApp.Notify()")
        global subjectCBDNumber
        global version
        global noProblem
        global bswIntegratorr
        global extIntegrator
        global pdfAttached
        global packageTypeFound

        if test_mode == 1:
            for i in range (len(bswIntegrators)):
                bswIntegrators[i] = test_name
                fblIntegrators[i] = test_name
                extIntegrators[i] = test_mail

        bswIntegratorsEmails, noneCorrectBswIntegratorName = Vector_IssueApp.BuildIntegratorEmail(bswIntegrators)
        test_log("BSW Integrators Emails: ", bswIntegratorsEmails)

        fblIntegratorsEmails, noneCorrectFblIntegratorName = Vector_IssueApp.BuildIntegratorEmail(fblIntegrators)
        test_log("FBL Integrators Emails: ", fblIntegratorsEmails)

        # send notification Emails
        for o in range(len(bswIntegratorsEmails)):
            if projectNumbers[o] == "":
                projectNumbers[o] = c_noProjectNumber
            #print(packageTypeFound)    
            # if packageTypeFound[o] == c_typeFblPackage and noneCorrectFblIntegratorName[o] == 0:
            #     #print(packageTypeFound[o])
            #     # send notification Email to FBL Integrator

            #     # outl = win32com.client.Dispatch(c_outlook)
            #     # notificationEmail = outl.CreateItem(0)
            #     # notificationEmail.To = fblIntegratorsEmails[o]
            #     #notificationEmail.To = 'Markus.Rigling@marquardt.com'

            #     print("Run -> o: %s", o)

            #     if bswIntegratorsEmails[o] != "":
            #         notificationEmail.CC = bswIntegratorsEmails[o]
            #         #notificationEmail.CC = 'Markus.Rigling@marquardt.com'
            #         print("%s - Email will be sent to %s with %s in CC" % (projectNumbers[o], fblIntegrators[o], bswIntegratorsEmails[o]))
            #         logging.info("%s - Email will be sent to %s with %s in CC" % (projectNumbers[o], fblIntegrators[o], bswIntegratorsEmails[o]))
            #     elif extIntegrators[o] != "" and extIntegrators[o] != None:
            #         notificationEmail.CC = extIntegrators[o]
            #         #notificationEmail.CC = 'Markus.Rigling@marquardt.com'
            #         print("%s - Email will be sent to %s with %s in CC" % (projectNumbers[o], fblIntegrators[o], extIntegrators[o]))
            #         logging.info("%s - Email will be sent to %s with %s in CC" % (projectNumbers[o], fblIntegrators[o], extIntegrators[o]))

            #     notificationEmail.Subject = c_subjectEmail % (projectNumbers[o], subjectCBDNumber, subjectDate)

            #     notificationEmail.Body = c_bodyEmailIntern % (c_pathIssueReports % (subjectCBDNumber, version), subjectCBDNumber, version, subjectDate)
            #     notificationEmail.SentOnBehalfOfName = 'VectorIssueReports'
            #     if send_Mails == 1:
            #         notificationEmail.Send()
            # elif extIntegrators[o] != "" and extIntegrators[o] != None:
            #     if "@" in extIntegrators[o]:
            #         # send notification Email to Ext Integrator with BSW Integrator in CC
            #         # outl = win32com.client.Dispatch(c_outlook)
            #         # notificationEmail = outl.CreateItem(0)
            #         # notificationEmail.To = extIntegrators[o]
            #         #notificationEmail.To = 'Markus.Rigling@marquardt.com'

            #         if bswIntegratorsEmails[o] != "":
            #             notificationEmail.CC = bswIntegratorsEmails[o]
            #             #notificationEmail.CC = 'Markus.Rigling@marquardt.com'
            #             print("%s - Email will be sent to %s with %s in CC" % (projectNumbers[o], extIntegrators[o], bswIntegratorsEmails[o]))
            #             logging.info("%s - Email will be sent to %s with %s in CC" % (projectNumbers[o], extIntegrators[o], bswIntegratorsEmails[o]))
            #         else:
            #             print("%s - Email will be sent to %s" % (projectNumbers[o], extIntegrators[o]))
            #             logging.info("%s - Email will be sent to %s" % (projectNumbers[o], extIntegrators[o]))

            #         notificationEmail.Subject = c_subjectEmail % (projectNumbers[o], subjectCBDNumber, subjectDate)

            #         if noProblem == 1:
            #             notificationEmail.Body = c_bodyEmailIntern % (c_pathIssueReports % (subjectCBDNumber, version), subjectCBDNumber, version, subjectDate)
            #         elif noProblem == 0:
            #             notificationEmail.Body = c_bodyEmailProblem % (subjectCBDNumber, version, subjectDate)

            #         try:
            #             notificationEmail.Attachments.Add(pdfAttached)
            #             notificationEmail.Attachments.Add(xmlAttached)
            #         except:
            #             pass
            #         notificationEmail.SentOnBehalfOfName = 'VectorIssueReports'
            #         if send_Mails == 1:
            #             notificationEmail.Send()
            #     else:
            #         # send notification Email to BSW Integrator
            #         if noneCorrectBswIntegratorName[o] == 0:
            #             print("%s - Email will be sent to %s" % (projectNumbers[o], bswIntegratorsEmails[o]))
            #             logging.info("%s - Email will be sent to %s" % (projectNumbers[o], bswIntegratorsEmails[o]))
            #             # outl = win32com.client.Dispatch(c_outlook)
            #             # notificationEmail = outl.CreateItem(0)
            #             # notificationEmail.To = bswIntegratorsEmails[o]
            #             #notificationEmail.To = 'Markus.Rigling@marquardt.com'
            #             notificationEmail.Subject = c_subjectEmail % (projectNumbers[o], subjectCBDNumber, subjectDate)

            #             notificationEmail.Body = c_bodyEmailIntern % (c_pathIssueReports % (subjectCBDNumber, version), subjectCBDNumber, version, subjectDate)
            #             notificationEmail.SentOnBehalfOfName = 'VectorIssueReports'
            #             if send_Mails == 1:
            #                 notificationEmail.Send()
            #         else:
            #             pass
            # else:
            #     # send notification Email to BSW Integrator
            #     if noneCorrectBswIntegratorName[o] == 0:
            #         print("%s - Email will be sent to %s" % (projectNumbers[o], bswIntegratorsEmails[o]))
            #         logging.info("%s - Email will be sent to %s" % (projectNumbers[o], bswIntegratorsEmails[o]))
            #         # outl = win32com.client.Dispatch(c_outlook)
            #         # notificationEmail = outl.CreateItem(0)
            #         # notificationEmail.To = bswIntegratorsEmails[o]
            #         # #notificationEmail.To = 'Markus.Rigling@marquardt.com'
            #         # notificationEmail.Subject = c_subjectEmail % (projectNumbers[o], subjectCBDNumber, subjectDate)

            #         # notificationEmail.Body = c_bodyEmailIntern % (c_pathIssueReports % (subjectCBDNumber, version), subjectCBDNumber, version, subjectDate)
            #         # notificationEmail.SentOnBehalfOfName = 'VectorIssueReports'
            #         # if send_Mails == 1:
            #         #     notificationEmail.Send()
            #     else:
            #         pass

        logging.info("EMAIL PROCESSED! \n")
        logging.debug("<-- Vector_IssueApp.Notify()")

    ########################################################################################################################
    # generate processed- and unprocessed issues XML
    ########################################################################################################################

    #def create_processed_issues_excel(processed_issues):
    #    path_processed_issues = os.path.dirname(xmlFileDir) + c_pathIssuesXlsx + '\\processed_issues.xlsx'
    #    df = pd.DataFrame(processed_issues, columns=[
    #        'Project Number', 'Subject CBD Number', 'Version', 'Subject Date', 'BSW Integrator', 'FBL Integrator', 'Ext Integrator', 'Email Sent'
    #    ])
    #    df.to_excel(path_processed_issues, index=False)
    #    print("Processed issues saved to 'processed_issues.xlsx'")


    # def create_processed_issues_excel(processed_issues):
    #     path_processed_issues = os.path.dirname(xmlFileDir) + c_pathIssuesXlsx + '\\processed_issues.xlsx'
        
    #     # Create DataFrame from new processed issues
    #     new_df = pd.DataFrame(processed_issues, columns=[
    #         'Project Number', 'Subject CBD Number', 'Version', 'Subject Date', 'BSW Integrator', 'FBL Integrator', 'Ext Integrator', 'Email Sent'
    #     ])
        
    #     if os.path.exists(path_processed_issues):
    #         # Read existing file
    #         existing_df = pd.read_excel(path_processed_issues)
    #         # Concatenate new data with existing data
    #         combined_df = pd.concat([existing_df, new_df], ignore_index=True)
    #         # Remove duplicates based on 'Project Number' and 'Subject CBD Number'
    #         combined_df.drop_duplicates(subset=['Project Number', 'Subject CBD Number', 'Version'], keep='last', inplace=True)
    #     else:
    #         # If file does not exist, use the new DataFrame
    #         combined_df = new_df
        
    #     # Save combined data to the Excel file
    #     combined_df.to_excel(path_processed_issues, index=False)
    #     print("Processed issues saved to 'processed_issues.xlsx'")

#    def create_unprocessed_issues_excel(unprocessed_issues):
#        path_unprocessed_issues = os.path.dirname(c_pathIssueReports) + c_pathUnprocessedIssues + '\\unprocessed_issues.xlsx'
#        df = pd.DataFrame(unprocessed_issues, columns=[
#            'Project Number', 'Subject CBD Number', 'Version', 'Subject Date', 'BSW Integrator', 'FBL Integrator', 'Ext Integrator', 'Reason'
#        ])
#        df.to_excel(path_unprocessed_issues, index=False)
#        print("Unprocessed issues saved to 'unprocessed_issues.xlsx'")

    def create_unprocessed_issues_excel(unprocessed_issues):
        path_unprocessed_issues = os.path.dirname(c_pathIssueReports) + c_pathUnprocessedIssues + '\\unprocessed_issues.xlsx'

        # Create DataFrame from new issues
        new_df = pd.DataFrame(unprocessed_issues, columns=[
            'Project Number', 'Subject CBD Number', 'Version', 'Subject Date', 'BSW Integrator', 'FBL Integrator', 'Ext Integrator', 'Reason'
        ])

        if os.path.exists(path_unprocessed_issues):
            # Read existing file
            existing_df = pd.read_excel(path_unprocessed_issues)
            # Concatenate new data with existing data
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            # If file does not exist, use the new DataFrame
            combined_df = new_df

        # Save combined data to the Excel file
        combined_df.to_excel(path_unprocessed_issues, index=False)
        print()
        print()
        print()
        print("Unprocessed issues saved to 'unprocessed_issues.xlsx'")

    ########################################################################################################################
    # send notification to cybersecurity manager
    ########################################################################################################################
    def send_summary_email(unprocessed_issues, unprocessed_mail):
        if unprocessed_issues and send_Mails == 1:
            # Format the email subject and body
            subject = "Summary mail, third party reports"
            body = Vector_IssueApp.format_summary_mail(unprocessed_issues, unprocessed_mail)

            # Use GraphAPI to send mails
            graph_access_certStore.send_mail(subject, body, c_cyberSecManager, c_issuesEmail)

    ########################################################################################################################
    # Build and Format the Body of the Notification Email for Unprocessed Issues
    ########################################################################################################################
    def format_unprocessed_issues(issues, unprocessed_mail):
        body = "Summary of Unprocessed Emails:\n\n"
        body += f"{'Sender':<35}|| {'Subject':<55}\n"
        body += f"{'CBD Number':<25}| {'Version':<10}| {'Date':<15}| {'Reason':<40}\n"
        body += "="*93 + "\n"  # Adds a separator for better readability

        for issue, mail in zip(issues, unprocessed_mail):
            cbd_number = issue[1] if issue[1] != 'Unknown' else "N/A"
            version = issue[2] if issue[2] != 'Unknown' else "N/A"
            subject_date = issue[3] if issue[3] != 'Unknown' else "N/A"
            reason = issue[7] if issue[7] != 'Unknown' else "Unknown Reason"
            subject, sender = mail

            # Format each line as: "CBD Number | Version | Date | Reason"
            body += f"{sender:<35}|| {subject:<55}\n"
            body += f"{cbd_number:<25}| {version:<10}| {subject_date:<15}| {reason:<40}\n"
            body += "-"*93 + "\n" 

        return body


    ########################################################################################################################
    # send notification to cybersecurity manager
    ########################################################################################################################
    def send_summary_Mail(non_standard_mail, epic_keys):
        if non_standard_mail and send_Mails == 1:
            # Format the email subject and body
            subject = "Summary mail, third party reports"
            body = Vector_IssueApp.format_summary_mail(non_standard_mail, epic_keys)

            # Use GraphAPI to send the email
            graph_access_certStore.send_mail(subject, body, [c_cyberSecManager, c_vulnManager], c_issuesEmail)


    ########################################################################################################################
    # Build and Format the Body of the Summary Email for third party issue reports
    ########################################################################################################################
    def format_summary_mail(non_standard_mail, epic_keys):
        body = "Summary of third party issue reports:\n\n"
        body += f"{'Sender':<35}|| {'Subject':<55}\n"
        body += f"{'Date':<35}|| {'epic_key':<55}\n"
        body += "="*93 + "\n"

        for mail, key in zip(non_standard_mail, epic_keys):
            #sender must be a list []
            sender = str(Vector_IssueApp.get_sender_email(mail)).strip()
            subject = str(mail['subject'])

            received_str = mail['receivedDateTime']
            received_dt = dt.datetime.strptime(received_str, "%Y-%m-%dT%H:%M:%SZ")
            subject_date = received_dt.strftime("%d.%m.%Y")
            epic_key = str(key)

            body += f"{sender:<30}|| {subject:<45}\n"
            body += f"{subject_date:<35}|| {epic_key:<55}\n"
            body += "-"*93 + "\n" 
        return body

if __name__ == "__main__":
        Vector_IssueApp.main()






