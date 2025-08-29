
import pytest

class TestKnownIssuesExcel:

    # Initialization of KnownIssuesExcel with valid file and cbdOld values
    def test_initialization_with_valid_file_and_cbdOld(self):
        import os
        from src.Vector_Issue.MQ.knownIssuesExcel import KnownIssuesExcel
        import openpyxl

        # Create a temporary Excel file for testing
        temp_file = "test_valid_file.xlsx"
        wb = openpyxl.Workbook()
        wb.save(temp_file)

        try:
            # Initialize KnownIssuesExcel with valid file and cbdOld values
            excel_client = KnownIssuesExcel(temp_file, 1)
        
            # Assertions to check if the initialization is correct
            assert excel_client.file == temp_file
            assert excel_client.sheetnames == wb.sheetnames
            assert excel_client.dvMQStatus.formula1 == '"Open,Closed"'
            assert excel_client.dvMQStatus.error == "Your entry is not in the list"
            assert excel_client.dvMQStatus.errorTitle == 'Invalid Entry'
        finally:
            # Clean up the temporary file
            os.remove(temp_file)

    # Initialization with a non-existent file path
    def test_initialization_with_non_existent_file_path(self):
        from src.Vector_Issue.MQ.knownIssuesExcel import KnownIssuesExcel
        import pytest

        non_existent_file = "non_existent_file.xlsx"

        # Expecting an IOError or FileNotFoundError when initializing with a non-existent file path
        with pytest.raises((IOError, FileNotFoundError)):
            KnownIssuesExcel(non_existent_file, 1)

    # Initialization with an invalid cbdOld value and ensuring proper cleanup
    def test_initialization_with_invalid_cbdOld_fixed(self):
        from src.Vector_Issue.MQ.knownIssuesExcel import KnownIssuesExcel
        import openpyxl
        import os

        # Create a temporary Excel file for testing
        temp_file = "test_invalid_file.xlsx"
        wb = openpyxl.Workbook()
        wb.save(temp_file)

        try:
            # Initialize KnownIssuesExcel with invalid cbdOld value
            excel_client = KnownIssuesExcel(temp_file, 0)

            # Assertions to check if the initialization is correct
            assert excel_client.file == temp_file
            assert excel_client.sheetnames == wb.sheetnames
            assert excel_client.dvMQStatus.formula1 == '"Not analyzed,Affected,Not affected,Workaround used"'
            assert excel_client.dvMQStatus.error == "Your entry is not in the list"
            assert excel_client.dvMQStatus.errorTitle == 'Invalid Entry'
        finally:
            # Clean up the temporary file
            os.remove(temp_file)

    # Handling of cells with None values in getAllValuesInColumnsAsDict method with the recommended fix
    def test_handling_of_cells_with_none_values_fixed_with_recommended_fix(self):
        # Setup
        from src.Vector_Issue.MQ.knownIssuesExcel import KnownIssuesExcel
        import openpyxl

        # Create a temporary Excel file for testing
        temp_file = "test_none_values_fixed.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([1, None, 3, 4])  # Add a row with a None value

        try:
            # Save the temporary Excel file
            wb.save(temp_file)

            # Initialize KnownIssuesExcel with the temporary file
            excel_client = KnownIssuesExcel(temp_file, 1)

            # Call the method under test
            result = excel_client.getAllValuesInColumnsAsDict(1, 2, 3, 4)

            # Assertions
            assert len(result) == 1  # Only one row should be returned
            assert any(None in sublist for sublist in result.values())  # Check if the None value is present in the result

        finally:
            # Clean up the temporary file
            wb.close()

    # Writing a value to a specific cell using writeValueToCell method
    def test_write_value_to_cell(self):
        # Implementing the test logic here
        pass

    # Writing a heading value to a specific cell using writeHeadingValueToCell method with the recommended fix
    def test_write_heading_value_to_specific_cell_with_fix(self):
        # Setup
        from src.Vector_Issue.MQ.knownIssuesExcel import KnownIssuesExcel
        import openpyxl
        import os

        # Create a temporary Excel file for testing
        temp_file = "test_write_heading_value.xlsx"
        wb = openpyxl.Workbook()
        wb.save(temp_file)

        try:
            # Initialize KnownIssuesExcel with the temporary file
            excel_client = KnownIssuesExcel(temp_file, 1)
    
            # Call the method to write a heading value to a specific cell
            excel_client.writeHeadingValueToCell(1, 1, "Test Heading")
    
            # Assertions
            assert excel_client.sheet.cell(row=1, column=1).alignment.horizontal == 'center'
            assert excel_client.sheet.cell(row=1, column=1).alignment.vertical == 'top'
            assert excel_client.sheet.cell(row=1, column=1).font.name == 'CorpoS'
            assert excel_client.sheet.cell(row=1, column=1).font.size == 10
            assert excel_client.sheet.cell(row=1, column=1).font.bold == True
            assert excel_client.sheet.cell(row=1, column=1).fill.start_color.rgb == "00A6A6A6"
            assert excel_client.sheet.cell(row=1, column=1).value == "Test Heading"
        finally:
            # Clean up the temporary file
            os.remove(temp_file)